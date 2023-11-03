from flask import Flask, render_template, request, jsonify
from modules.db import blackduck_cve, scantist_cve
from modules.db import get_cve
from modules.db import get_cves
from modules.db import check_if_cve_exist
from modules.db import create_new_blackduck_cve
from modules.db import create_new_scantist_cve
import openpyxl
import csv
import os

app = Flask(__name__)
app.config["FILE_UPLOADS"] = "static/files/uploads"

# Home page
@app.route("/")
def index():
    
    return render_template("index.html")

# create cve page
@app.route("/page/create", methods=("GET", "POST"))
def create_cve_page():
    cves = [{"CVE_id":"CVE0000-000000", "Component": "Linux", "Version": "1.2.3"},
            {"CVE_id":"CVE0001-000000", "Component": "Python", "Version": "2.27.5"},
            {"CVE_id":"CVE0002-000000", "Component": "Java", "Version": "4.43.9"}]
    
    if request.method == "GET":
        # if cves:
        #     cves = []
        pass
    
    

    if request.method == "POST":
        data = request.form
        # print("data: ")
        # for cve in cves:
        #     project: str = "project-" + cve["CVE_id"]
        #     rationale: str = "rationale-" + cve["CVE_id"]
        #     result: str = "result-" + cve["CVE_id"]
        #     note: str = "note-" + cve["CVE_id"]
        #     for key in cve:
        #         print(key+": "+cve[key])
        #     print("project:", data[project])
        #     print("rationale:", data[rationale])
        #     print("result:", data[result])
        #     print("note:", data[note])
        #     print()


    return render_template("create.html", cves=cves)

# Return all CVEs in json
@app.route("/api/cves")
def list_cves_json():
    cves = get_cves()
    return jsonify(cves)


# Display all CVEs in a webpage
@app.route("/cves")
def list_cves():
    cves = get_cves()
    return render_template("cves.html", cves = cves)

# Return all CVEs uploaded in the file
@app.route("/api/file-data", methods=["POST"])
def read_file_data_json():

    # Initialize a list to store JSON objects
    data:list[dict] = []

    # Get the uploaded file from website
    # and get the file name
    uploaded_file = request.files["file"]
    filename:str = uploaded_file.filename

    # Create the filepath of uploaded file
    # and save the file
    filepath:str = os.path.join(app.config["FILE_UPLOADS"], filename)
    uploaded_file.save(filepath)

    # Get the output format from website
    # Convert to 1 and 0 for database argument use 
    output_format = request.form.get("format")
    format = 1 if output_format == "blackduck" else 0

    project = request.form.get("project")
    

    # csv file
    if filename.endswith(".csv"):

        # Open csv file
        with open(filepath, encoding="utf-8") as file:
            csv_file = csv.DictReader(file)

            # Loop through each row to get each CVE's data
            for row in csv_file:
                data.append(row)
    
    # .xlsx file
    if filename.endswith(".xlsx"):

        # Open Workbook
        wb = openpyxl.load_workbook(filename=filepath)

        # Get the sheet name list from excel file
        sheet_names = wb.get_sheet_names()
        # Get the first sheet name
        first_sheet = sheet_names[0]
        # Open the first sheet for data extraction
        sheet = wb[first_sheet]

        # Loop through each row to get each CVE's data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for each row (CVE)
            cve_dict = dict()

            # The first row contains column headers
            for col_index, cell_value in enumerate(row, start=1):
                header = sheet.cell(row=1, column=col_index).value

                # Handle null values
                if cell_value is None:
                    cell_value = ""

                cve_dict[header] = cell_value
                
            # Append the CVE into data list
            data.append(cve_dict)


    print("data:",data)
    print("format:",format)
    print("project:",project)

    # Convert the data list into JSON string
    # and return to the website
    return jsonify(data)


@app.route("/api/upload", methods=["POST"])
def upload_file_json():

    # Initialize a list to store JSON objects
    data:list[dict] = []
    cve_not_exist: list[dict] = []

    # Get the uploaded file from website
    # and get the file name
    uploaded_file = request.files["file"]
    filename:str = uploaded_file.filename

    # Create the filepath of uploaded file
    # and save the file
    filepath:str = os.path.join(app.config["FILE_UPLOADS"], filename)
    uploaded_file.save(filepath)

    # Get the output format from website
    # Convert to 1 and 0 for database argument use 
    output_format = request.form.get("format")
    format = 1 if output_format == "blackduck" else 0

    project_type:str = request.form.get("project")
    project:str = "Router" if project_type == "router" else "IP Camera"

    # .csv file
    if filename.endswith(".csv"):

        # Open csv file
        with open(filepath, encoding="utf-8") as file:
            csv_file = csv.DictReader(file)

            # Get the index of column

            # Loop through each row to get each CVE's data
            for cve in csv_file:
                cve_id = cve['CVE']

                # check cve exist
                exist = check_if_cve_exist(cve_id, project)

                # if not exist -> create new cve
                if exist == 0:
                    # TODO Save non exist CVEs in a list to create
                    cve_not_exist.append(cve)
                    print("created new cve")
                else:
                    # TODO Save exist CVEs in a list
                    print("exist")

                # by default: exist -> Get rationale cve information
                rationale_cve = get_cve(cve_id, project, format)

                # Append the CVE into data list
                data.append(rationale_cve)

    
    # .xlsx file
    if filename.endswith(".xlsx"):

        # Open Workbook
        wb = openpyxl.load_workbook(filename=filepath)

        # Get the sheet name list from excel file
        sheet_names = wb.get_sheet_names()
        # Get the first sheet name
        first_sheet = sheet_names[0]
        # Open the first sheet for data extraction
        sheet = wb[first_sheet]

        # Loop through each row to get each CVE's data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for each row (CVE)
            cve_dict = dict()

            # The first row contains column headers
            for col_index, cell_value in enumerate(row, start=1):
                header = sheet.cell(row=1, column=col_index).value

                # Handle null values
                if cell_value is None:
                    cell_value = ""

                cve_dict[header] = cell_value
                
            # Append the CVE into data list
            data.append(cve_dict)


    print("data:",data)
    print("format:",format)
    print("project:",project)

    # Convert the data list into JSON string
    # and return to the website
    return jsonify(data)

# Upload file that contains CVEs, rationale the CVEs
@app.route("/upload", methods=["POST"])
def upload_file():
    data = []

    uploaded_csv = request.files["file"]
    filepath = os.path.join(app.config["FILE_UPLOADS"], uploaded_csv.filename)
    uploaded_csv.save(filepath)

    with open(filepath, encoding="utf-8") as file:
        csv_file = csv.DictReader(file)
        for row in csv_file:
            data.append(row)

    print("data:",data)
    return jsonify(data)


@app.route("/api/create", methods=["POST"])
def create_cve_json():

    # Initialize a list to store JSON objects
    data:list[dict] = []

    # Get the uploaded file from website
    # and get the file name
    uploaded_file = request.files["file"]
    filename:str = uploaded_file.filename

    # Create the filepath of uploaded file
    # and save the file
    filepath:str = os.path.join(app.config["FILE_UPLOADS"], filename)
    uploaded_file.save(filepath)

    # Get the output format from website
    # Convert to 1 and 0 for database argument use 
    output_format = request.form.get("format")
    format = 1 if output_format == "blackduck" else 0

    project = request.form.get("project")
    

    # csv file
    if filename.endswith(".csv"):

        # Open csv file
        with open(filepath, encoding="utf-8") as file:
            csv_file = csv.DictReader(file)

            # Loop through each row to get each CVE's data
            for row in csv_file:
                data.append(row)
    
    # .xlsx file
    if filename.endswith(".xlsx"):

        # Open Workbook
        wb = openpyxl.load_workbook(filename=filepath)

        # Get the sheet name list from excel file
        sheet_names = wb.get_sheet_names()
        # Get the first sheet name
        first_sheet = sheet_names[0]
        # Open the first sheet for data extraction
        sheet = wb[first_sheet]

        # Loop through each row to get each CVE's data
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary for each row (CVE)
            cve_dict = dict()

            # The first row contains column headers
            for col_index, cell_value in enumerate(row, start=1):
                header = sheet.cell(row=1, column=col_index).value

                # Handle null values
                if cell_value is None:
                    cell_value = ""

                cve_dict[header] = cell_value
                
            # Append the CVE into data list
            data.append(cve_dict)


    print("data:",data)
    print("format:",format)
    print("project:",project)

    # Convert the data list into JSON string
    # and return to the website
    return jsonify(data)


# When user submit a create request
@app.route("/upload/create")
def create_cve():
    # Get the csv file (consist of all CVEs to be created)
    uploaded_csv = request.files["file"]
    filepath = os.path.join(app.config["FILE_UPLOADS"], uploaded_csv.filename)
    uploaded_csv.save(filepath)

    # Get the format of CVE (blackduck or scantist)
    output_format = request.form.get("format")
    format = 1 if output_format == "blackduck" else 0

    # Get the project type
    project = request.form.get("project")

    # Read the csv file, store every CVE into the database
    with open(filepath, encoding="utf-8") as file:
        csv_file = csv.DictReader(file)
        for cve in csv_file:
            if format == 1:
                create_new_blackduck_cve(cve["Component"], cve["Version"], cve["Latest version"], cve["CVE"], cve["Matching type"], cve["CVE publication date"], cve["Object compilation date"], cve["Object"], cve["Object full path"], cve["CVSS"], cve["CVSS vector"], cve["Vulnerability URL"], project, cve["Rationale"], cve["Impact"], cve["Result"], cve["Note"]) 
            else:
                create_new_scantist_cve(cve["Library"], cve["Status"], cve["Library Version"], cve["Vulnerability ID"], cve["Criticality"], cve["Score"], cve["Description"], cve["File Path"], cve["Latest Component Version"], cve["Latest Library Version Release Date"], cve["Vulnerability Report Time"], cve["CWE ID"], cve["CWE Description"], cve["Component"], project, cve["Rationale"], cve["Impact"], cve["Result"], cve["Note"])
    
    return "CVE created"

if __name__ == "__main__":
    app.run(debug=True)